#!/usr/bin/env python3
"""
update_data.py — обновляет TS-файлы данных рейтинга ПКО-300
из свежей выгрузки в data/source/PKO-*.xlsx.

Использование:
    python3 scripts/update_data.py --dry-run    # показывает diff и топ-20, не пишет файлы
    python3 scripts/update_data.py              # реальное обновление

Что делает:
    1. Находит самый свежий PKO-*.xlsx в data/source/ (или берёт --source).
    2. Парсит xlsx (openpyxl).
    3. Читает текущие ts-файлы для сохранения ручных правок:
       - ratingData.ts: napka, capitalAttraction
       - companyDetails.ts: website, fundraising, bonds
    4. Сортирует компании по полю «Выручка + прочие доходы 2025» (по убыванию).
    5. Присваивает плотный ранг 1..N.
    6. Если «Краткое название» пусто — fallback на полное через stripOrgForm.
    7. Перезаписывает ratingData.ts, financeDynamic.ts, companyDetails.ts.
    8. Печатает лог: добавлено/удалено/перемещено + топ-20.

Заменяет старые скрипты:
    scripts/convert_new_json.py  (читал JSON из Downloads, ранг из источника)
    scripts/prune_companies.py   (резал по keep-списку)
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO = Path(__file__).resolve().parent.parent
SOURCE_DIR = REPO / "data" / "source"
DATA_DIR = REPO / "design" / "src" / "app" / "data"
RATING_TS = DATA_DIR / "ratingData.ts"
FINANCE_TS = DATA_DIR / "financeDynamic.ts"
DETAILS_TS = DATA_DIR / "companyDetails.ts"

YEARS = [2021, 2022, 2023, 2024, 2025]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def parse_num(val) -> float:
    """'22 501 574,00' → 22501574.0; None/'nan'/'—' → 0.0"""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val) if val == val else 0.0  # NaN check
    s = str(val).strip()
    if not s or s in ("—", "-") or s.lower() in ("nan", "inf"):
        return 0.0
    s = s.replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


ORGFORM_FULL_TO_SHORT = [
    (r"^Непубличное\s+акционерное\s+общество\s*", "НАО"),
    (r"^Публичное\s+акционерное\s+общество\s*", "ПАО"),
    (r"^Закрытое\s+акционерное\s+общество\s*", "ЗАО"),
    (r"^Открытое\s+акционерное\s+общество\s*", "ОАО"),
    (r"^Акционерное\s+общество\s*", "АО"),
    (r"^Общество\s+с\s+ограниченной\s+ответственностью\s*", "ООО"),
]


def build_short_from_full(full: str) -> str:
    """Из полного юр. названия собирает каноничное краткое 'ОФ ПКО "Бренд"'.

    Пример:
        'Общество с ограниченной ответственностью "ПРОФЕССИОНАЛЬНАЯ КОЛЛЕКТОРСКАЯ
         ОРГАНИЗАЦИЯ "ПКО-АВЗ"' → 'ООО ПКО "ПКО-АВЗ"'
    """
    if not full:
        return ""
    s = full.strip()
    # 1. Определить + заменить organizational form
    orgform = ""
    for pat, abbr in ORGFORM_FULL_TO_SHORT:
        new_s, n = re.subn(pat, "", s, count=1, flags=re.IGNORECASE)
        if n > 0:
            orgform = abbr
            s = new_s
            break
    if not orgform:
        # Уже сокращённая форма (НАО/ООО/АО/...) — оставим как есть
        m = re.match(r"^(НАО|ООО|ПАО|АО|ЗАО|ОАО)\s+", s, flags=re.IGNORECASE)
        if m:
            orgform = m.group(1).upper()
            s = s[m.end():]
    # 2. Очистить от ведущих кавычек и пробелов
    s = re.sub(r'^["\s]+', "", s)
    # 3. Убрать "ПРОФЕССИОНАЛЬНАЯ КОЛЛЕКТОРСКАЯ ОРГАНИЗАЦИЯ" (полное) → ПКО
    pko_marker = ""
    new_s, n = re.subn(
        r'^"?ПРОФЕССИОНАЛЬНАЯ\s+КОЛЛЕКТОРСКАЯ\s+ОРГАНИЗАЦИЯ"?\s*',
        "",
        s,
        count=1,
        flags=re.IGNORECASE,
    )
    if n > 0:
        pko_marker = "ПКО"
        s = new_s
    else:
        new_s, n = re.subn(r'^"?ПКО"?\s*', "", s, count=1, flags=re.IGNORECASE)
        if n > 0:
            pko_marker = "ПКО"
            s = new_s
    # 4. Очистить остаток
    s = re.sub(r'^["\s]+', "", s)
    s = re.sub(r'["\s]+$', "", s)
    # 5. Сборка
    parts = [p for p in [orgform, pko_marker] if p]
    if s:
        parts.append(f'"{s}"')
    return " ".join(parts)


def strip_org_form(name: str) -> str:
    """Совместимость: возвращает только бренд без orgform/ПКО (для display-логики)."""
    short = build_short_from_full(name)
    s = re.sub(r'^(НАО|ООО|ПАО|АО|ЗАО|ОАО)\s+', "", short)
    s = re.sub(r'^ПКО\s+', "", s)
    return s.strip('"').strip()


def short_name(row: dict) -> str:
    """Краткое название из источника, fallback на полное через build_short_from_full."""
    short = (row.get("Краткое название") or "").strip()
    if short and short.lower() != "nan":
        return short
    full = (row.get("Полное название") or row.get("Название") or "").strip()
    return build_short_from_full(full) if full else ""


def clean_city(region: str) -> str:
    """'Г.Москва' → 'Москва'"""
    if not region:
        return ""
    r = region.strip()
    if r.lower().startswith("г."):
        return r[2:]
    return r


def fmt_date(raw: str) -> str:
    """'03.02.2009' → '2009-02-03'"""
    if not raw:
        return ""
    parts = str(raw).split(".")
    if len(parts) == 3:
        return f"{parts[2]}-{parts[1]}-{parts[0]}"
    return ""


# ---------------------------------------------------------------------------
# Source loading
# ---------------------------------------------------------------------------
def find_latest_source() -> Path:
    files = sorted(SOURCE_DIR.glob("PKO-*.xlsx"))
    if not files:
        sys.exit(f"ERROR: no PKO-*.xlsx files in {SOURCE_DIR}")
    return files[-1]


def load_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers, row))
        d["__inn"] = str(d.get("ИНН", "")).strip()
        rows.append(d)
    return rows


# ---------------------------------------------------------------------------
# Read existing TS data (для сохранения ручных правок)
# ---------------------------------------------------------------------------
def read_existing_rating() -> dict[str, dict]:
    """ratingData.ts → {inn: {napka, capitalAttraction, name}}"""
    if not RATING_TS.exists():
        return {}
    text = RATING_TS.read_text(encoding="utf-8")
    result = {}
    # каждая запись на одной строке вида {"rank":...,"inn":"123",...},
    for line in text.splitlines():
        line = line.strip().rstrip(",")
        if not line.startswith("{") or not line.endswith("}"):
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            continue
        inn = obj.get("inn")
        if not inn:
            continue
        result[inn] = {
            "napka": obj.get("napka", False),
            "capitalAttraction": obj.get("capitalAttraction", "none"),
            "name": obj.get("name", ""),
        }
    return result


def read_existing_details() -> dict[str, dict]:
    """companyDetails.ts → {inn: {website, fundraising, bonds}}"""
    if not DETAILS_TS.exists():
        return {}
    text = DETAILS_TS.read_text(encoding="utf-8")
    result = {}
    # формат: "INN": { ... },
    for line in text.splitlines():
        m = re.match(r'\s*"(\d+)"\s*:\s*(\{.*\})\s*,?\s*$', line)
        if not m:
            continue
        inn = m.group(1)
        try:
            obj = json.loads(m.group(2))
        except json.JSONDecodeError:
            continue
        result[inn] = {
            "website": obj.get("website", ""),
            "fundraising": obj.get("fundraising"),
            "bonds": obj.get("bonds"),
        }
    return result


# ---------------------------------------------------------------------------
# TS generation
# ---------------------------------------------------------------------------
RATING_HEADER = '''// Auto-generated by scripts/update_data.py — НЕ РЕДАКТИРОВАТЬ ВРУЧНУЮ.
// Источник: data/source/{src}
// Ранжирование: по полю «Выручка + прочие доходы 2025» (по убыванию).

export interface RatingCompany {{
  rank: number;
  name: string;
  inn: string;
  city: string;       // Город регистрации
  revenue: number;    // Совокупный доход 2025 (выручка + прочие), тыс руб
  profit: number;     // Чистая прибыль 2025, тыс руб
  yearChange: number; // Δ к 2024 (%)
  growthRate: number; // Рост фин активов за 5 лет (%)
  experience: number; // Стаж (лет)
  capitalAttraction: 'public' | 'corporate' | 'none';
  napka: boolean;
  cost: number;        // Расходы 2025, тыс руб
  eqt: number;         // Собственный капитал, тыс руб
  dLong: number;       // Долгосрочные обязательства, тыс руб
  dShort: number;      // Краткосрочные обязательства, тыс руб
  de: number;          // D/E коэффициент
  receivable: number;  // Дебиторская задолженность, тыс руб
  cagr: number;        // CAGR выручки 5 лет (доля)
  loan: number;        // Заёмный капитал, тыс руб
  rankDelta: number;   // Изменение позиции в рейтинге YoY
}}

export const ratingData: RatingCompany[] = [
'''


FINANCE_HEADER = '''// Auto-generated by scripts/update_data.py — НЕ РЕДАКТИРОВАТЬ ВРУЧНУЮ.
// Источник: data/source/{src}
// Индексы массивов: [0]=2021, [1]=2022, [2]=2023, [3]=2024, [4]=2025

export const FINANCE_YEARS = [2021, 2022, 2023, 2024, 2025] as const;
export type FinanceYearType = typeof FINANCE_YEARS[number];

export interface FinanceDynamic {{
  rank: number;
  name: string;
  inn: string;
  income: number[];     // Выручка + пр. доходы, тыс руб
  cost: number[];       // Расходы, тыс руб
  profit: number[];     // Чистая прибыль, тыс руб
  receivable: number[]; // Дебиторская задолженность, тыс руб
}}

export const financeDynamic: FinanceDynamic[] = [
'''


DETAILS_HEADER = '''// Auto-generated by scripts/update_data.py — НЕ РЕДАКТИРОВАТЬ ВРУЧНУЮ.
// Источник: data/source/{src}

export interface YearlyFinancials {{
  year: number;
  revenue: number;
  cost: number;
  profit: number;
  assets: number;
  receivable: number;
}}

export interface CapitalStructure {{
  equity: number;
  debt: number;
  deRatio: number;
  authorizedCapital: number;
  debtShare: number;
  equityShare: number;
  totalAssets: number;
}}

export interface FundraisingInfo {{
  type: string;
  status: string;
  details: string;
  founder: string;
}}

export interface BondInfo {{
  rating: string;
  isin: string;
  volume: number;
  coupon: string;
  maturity: string;
  status: string;
  ticker: string;
}}

export interface CompanyDetails {{
  inn: string;
  fullName: string;
  director: string;
  ogrn: string;
  registrationDate: string;
  website: string;
  region: string;
  address: string;
  authorizedCapital: number;
  revenue2025: number;
  otherIncome2025: number;
  revenue2024: number;
  otherIncome2024: number;
  financials: YearlyFinancials[];
  capitalStructure: CapitalStructure;
  fundraising: FundraisingInfo | null;
  bonds: BondInfo[] | null;
}}

export const companyDetailsMap: Record<string, CompanyDetails> = {{
'''


def build_rating_entry(row: dict, rank: int, name: str, old_rating: dict, rank_2024: dict) -> dict:
    inn = row["__inn"]
    rev_2025 = parse_num(row.get("Выручка + прочие доходы 2025"))
    rev_2024 = parse_num(row.get("Выручка + прочие доходы 2024"))
    year_change = round(((rev_2025 - rev_2024) / rev_2024) * 100, 1) if rev_2024 > 0 else 0.0
    # rankDelta: позиция-2024 минус позиция-2025 в НАШЕМ ранжировании (по «Выручка + прочие доходы»).
    # Колонку источника «Позиция относительно прошлого года» не используем — она от «Номера рейтинга ПКО-300».
    prev_rank = rank_2024.get(inn)
    rank_delta = (prev_rank - rank) if prev_rank else 0
    de_pct = parse_num(row.get("D/E коэффициент"))
    cagr_pct = parse_num(row.get("Темпы роста за 5 лет (CAGR)"))
    saved = old_rating.get(inn, {})
    return {
        "rank": rank,
        "name": name,
        "inn": inn,
        "city": clean_city(row.get("Регион", "")),
        "revenue": round(rev_2025),
        "profit": round(parse_num(row.get("Чистая прибыль 2025"))),
        "yearChange": year_change,
        "growthRate": round(parse_num(row.get("Рост фин активов за 5 лет")), 1),
        "experience": int(parse_num(row.get("Стаж", 0))),
        "capitalAttraction": saved.get("capitalAttraction", "none"),
        "napka": saved.get("napka", False),
        "cost": round(parse_num(row.get("Расходы 2025"))),
        "eqt": round(parse_num(row.get("Собственный капитал 2025"))),
        "dLong": 0,
        "dShort": 0,
        "de": round(de_pct, 2) if de_pct else 0.0,
        "receivable": round(parse_num(row.get("Дебит. задолженность 2025"))),
        "cagr": round(cagr_pct / 100, 2) if cagr_pct else 0.0,
        "loan": round(parse_num(row.get("Заёмный капитал 2025"))),
        "rankDelta": rank_delta,
    }


def build_finance_entry(row: dict, rank: int, name: str) -> dict:
    return {
        "rank": rank,
        "name": name,
        "inn": row["__inn"],
        "income": [round(parse_num(row.get(f"Выручка + прочие доходы {y}"))) for y in YEARS],
        "cost": [round(parse_num(row.get(f"Расходы {y}"))) for y in YEARS],
        "profit": [round(parse_num(row.get(f"Чистая прибыль {y}"))) for y in YEARS],
        "receivable": [round(parse_num(row.get(f"Дебит. задолженность {y}"))) for y in YEARS],
    }


def build_details_entry(row: dict, old_details: dict) -> dict:
    inn = row["__inn"]
    saved = old_details.get(inn, {})
    financials = []
    for y in YEARS:
        financials.append({
            "year": y,
            "revenue": round(parse_num(row.get(f"Выручка {y}"))),
            "cost": round(parse_num(row.get(f"Расходы {y}"))),
            "profit": round(parse_num(row.get(f"Чистая прибыль {y}"))),
            "assets": round(parse_num(row.get(f"Фин. вложения {y}"))),
            "receivable": round(parse_num(row.get(f"Дебит. задолженность {y}"))),
        })
    eq = parse_num(row.get("Собственный капитал 2025"))
    debt = parse_num(row.get("Заёмный капитал 2025"))
    total = eq + debt
    return {
        "inn": inn,
        "fullName": row.get("Полное название") or row.get("Название") or "",
        "director": row.get("Генеральный директор") or "",
        "ogrn": str(row.get("ОГРН") or ""),
        "registrationDate": fmt_date(row.get("Дата основания") or ""),
        "website": saved.get("website", row.get("Сайт") or ""),
        "region": row.get("Регион") or "",
        "address": row.get("Адрес") or "",
        "authorizedCapital": round(parse_num(row.get("Уставной капитал"))),
        "revenue2025": round(parse_num(row.get("Выручка 2025"))),
        "otherIncome2025": round(parse_num(row.get("Прочие доходы 2025"))),
        "revenue2024": round(parse_num(row.get("Выручка 2024"))),
        "otherIncome2024": round(parse_num(row.get("Прочие доходы 2024"))),
        "financials": financials,
        "capitalStructure": {
            "equity": round(eq),
            "debt": round(debt),
            "deRatio": round(parse_num(row.get("D/E коэффициент")), 2),
            "authorizedCapital": round(parse_num(row.get("Уставной капитал"))),
            "debtShare": round((debt / total) * 100, 1) if total > 0 else 0.0,
            "equityShare": round((eq / total) * 100, 1) if total > 0 else 0.0,
            "totalAssets": round(total),
        },
        "fundraising": saved.get("fundraising"),
        "bonds": saved.get("bonds"),
    }


def write_ts_files(rows_with_rank: list[tuple[int, str, dict]], src_name: str,
                   old_rating: dict, old_details: dict, rank_2024: dict):
    # ratingData.ts
    out = RATING_HEADER.format(src=src_name)
    for rank, name, row in rows_with_rank:
        e = build_rating_entry(row, rank, name, old_rating, rank_2024)
        out += f"  {json.dumps(e, ensure_ascii=False)},\n"
    out += "];\n"
    RATING_TS.write_text(out, encoding="utf-8")

    # financeDynamic.ts
    out = FINANCE_HEADER.format(src=src_name)
    for rank, name, row in rows_with_rank:
        e = build_finance_entry(row, rank, name)
        out += f"  {json.dumps(e, ensure_ascii=False)},\n"
    out += "];\n"
    FINANCE_TS.write_text(out, encoding="utf-8")

    # companyDetails.ts
    out = DETAILS_HEADER.format(src=src_name)
    for rank, name, row in rows_with_rank:
        e = build_details_entry(row, old_details)
        out += f'  "{e["inn"]}": {json.dumps(e, ensure_ascii=False)},\n'
    out += "};\n"
    DETAILS_TS.write_text(out, encoding="utf-8")


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def validate(rows_with_rank: list[tuple[int, str, dict]]) -> list[str]:
    errors = []
    inns = [row["__inn"] for _, _, row in rows_with_rank]
    if len(set(inns)) != len(inns):
        errors.append("ИНН не уникальны")
    ranks = [r for r, _, _ in rows_with_rank]
    expected = list(range(1, len(rows_with_rank) + 1))
    if ranks != expected:
        errors.append(f"Ранги не плотные 1..N (max={max(ranks)}, count={len(ranks)})")
    for _, name, row in rows_with_rank:
        if not name:
            errors.append(f"Пустое имя для ИНН {row['__inn']}")
        if not row["__inn"]:
            errors.append("Пустой ИНН")
    return errors


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--dry-run", action="store_true", help="Не писать файлы, только показать diff")
    parser.add_argument("--source", help="Путь к xlsx (по умолчанию — самый свежий в data/source/)")
    args = parser.parse_args()

    src = Path(args.source) if args.source else find_latest_source()
    print(f"📂 Source: {src.relative_to(REPO)}")

    rows = load_xlsx(src)
    print(f"📊 Loaded {len(rows)} rows from xlsx")

    # Dedupe by INN
    seen = set()
    uniq = []
    for r in rows:
        inn = r["__inn"]
        if not inn or inn in seen:
            continue
        seen.add(inn)
        uniq.append(r)
    if len(uniq) < len(rows):
        print(f"⚠️  Removed {len(rows) - len(uniq)} duplicate INNs")
    rows = uniq

    # Sort by «Выручка + прочие доходы 2025», desc
    rows.sort(key=lambda r: parse_num(r.get("Выручка + прочие доходы 2025")), reverse=True)

    # Plain rank 1..N + computed name
    rows_with_rank = []
    for i, r in enumerate(rows, start=1):
        name = short_name(r)
        rows_with_rank.append((i, name, r))

    # Ранжирование за 2024 по той же метрике — для расчёта rankDelta.
    rows_2024 = sorted(
        [r for r in rows if parse_num(r.get("Выручка + прочие доходы 2024")) > 0],
        key=lambda r: parse_num(r.get("Выручка + прочие доходы 2024")), reverse=True)
    rank_2024 = {r["__inn"]: i for i, r in enumerate(rows_2024, start=1)}

    # Validate
    errors = validate(rows_with_rank)
    if errors:
        print("❌ Validation failed:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    print("✅ Validation passed")

    # Read existing TS for diff
    old_rating = read_existing_rating()
    old_details = read_existing_details()

    new_inns = {r["__inn"] for r in rows}
    old_inns = set(old_rating.keys())

    print()
    print("=" * 70)
    print("📈 DIFF vs current ratingData.ts")
    print("=" * 70)
    print(f"  Total: {len(rows)} (was {len(old_rating)})")
    print(f"  Added: {len(new_inns - old_inns)}")
    print(f"  Removed: {len(old_inns - new_inns)}")
    print(f"  Kept: {len(new_inns & old_inns)}")

    added = new_inns - old_inns
    if added:
        print(f"\n  Added INNs:")
        for r in rows:
            if r["__inn"] in added:
                print(f"    {r['__inn']:>12} | {short_name(r)[:50]}")

    removed = old_inns - new_inns
    if removed:
        print(f"\n  Removed INNs:")
        for inn in sorted(removed):
            print(f"    {inn:>12} | {old_rating[inn].get('name', '?')[:50]}")

    # Top-20
    print()
    print("=" * 70)
    print("🏆 TOP-20 (по «Выручка + прочие доходы 2025»)")
    print("=" * 70)
    print(f"  {'#':>3} | {'ИНН':>12} | {'Название':<35} | {'Total ₽млн':>10} | {'Выручка':>10} | {'Прочие':>10}")
    print("  " + "-" * 100)
    for rank, name, row in rows_with_rank[:20]:
        ti = parse_num(row.get("Выручка + прочие доходы 2025")) / 1000  # тыс → млн
        rv = parse_num(row.get("Выручка 2025")) / 1000
        oi = parse_num(row.get("Прочие доходы 2025")) / 1000
        marker = "🆕 " if row["__inn"] in added else "  "
        print(f"  {marker}№{rank:>3} | {row['__inn']:>12} | {name[:35]:<35} | {ti:>10,.0f} | {rv:>10,.0f} | {oi:>10,.0f}")

    if args.dry_run:
        print()
        print("🔬 DRY-RUN: no files written. Re-run without --dry-run to apply.")
        return

    # Real write
    src_name = src.name
    write_ts_files(rows_with_rank, src_name, old_rating, old_details, rank_2024)
    print()
    print(f"✅ Wrote {len(rows)} entries → {RATING_TS.name}, {FINANCE_TS.name}, {DETAILS_TS.name}")


if __name__ == "__main__":
    main()
