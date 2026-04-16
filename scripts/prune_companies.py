#!/usr/bin/env python3
"""
Prune the company database down to the 529 active PKOs from
/Users/valeriapinaeva/Downloads/PKO (2).xlsx (sheet "ПКО").

Removes 104 liquidated companies from:
  - design/src/app/data/ratingData.ts      (array, reranks 1..N)
  - design/src/app/data/financeDynamic.ts  (array, reranks 1..N)
  - design/src/app/data/companyDetails.ts  (Record<INN,...>)
  - design/src/app/data/logoMap.ts         (Record<INN, filename>)

Verifies (but does not mutate) investmentData.ts — aborts if any
removed company is still referenced there by short name.

Idempotent: second run on already-pruned files makes no changes.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO = Path(__file__).resolve().parent.parent
SOURCE_JSON = Path("/Users/valeriapinaeva/Downloads/PKO (2).json")
SOURCE_XLSX = Path("/Users/valeriapinaeva/Downloads/PKO (2).xlsx")
DATA = REPO / "design" / "src" / "app" / "data"

RATING_TS = DATA / "ratingData.ts"
FINANCE_TS = DATA / "financeDynamic.ts"
DETAILS_TS = DATA / "companyDetails.ts"
LOGO_TS = DATA / "logoMap.ts"
INVEST_TS = DATA / "investmentData.ts"

# ---------------------------------------------------------------------------
# 1. Read keep-list + full names from JSON (preferred) or Excel (fallback)
# ---------------------------------------------------------------------------
def load_keep_data() -> tuple[set[str], dict[str, str]]:
    """Return (keep_inns, inn_to_search_text).

    search_text = fullName + " | " + shortName, used for substring matching
    against investmentData brand names (e.g. «ПКБ», «СКМ» appear only in short).
    """
    if SOURCE_JSON.exists():
        rows = json.loads(SOURCE_JSON.read_text(encoding="utf-8"))
        inns: set[str] = set()
        full: dict[str, str] = {}
        for r in rows:
            inn = r.get("ИНН")
            if inn is None:
                continue
            inn_s = str(inn).strip()
            inns.add(inn_s)
            parts = [
                str(r.get("Полное название") or ""),
                str(r.get("Название") or ""),
                str(r.get("Краткое название") or ""),
            ]
            full[inn_s] = " | ".join(p for p in parts if p).strip()
        return inns, full

    # Fallback to Excel
    import openpyxl

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    if "ПКО" not in wb.sheetnames:
        sys.exit(f"Excel sheet 'ПКО' not found. Sheets: {wb.sheetnames}")
    ws = wb["ПКО"]
    inns = set()
    full = {}
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h or "") for h in row]
            continue
        if row[0] is None:
            continue
        inn_s = str(row[0]).strip()
        inns.add(inn_s)
        idx = headers.index("Полное название") if "Полное название" in headers else 3
        full[inn_s] = str(row[idx] or row[1] or "").strip()
    return inns, full


# ---------------------------------------------------------------------------
# Helpers for line-based JSON-per-line arrays
# ---------------------------------------------------------------------------
OBJECT_LINE_RE = re.compile(r"^\s*(\{.*\}),?\s*$")


def filter_line_array(path: Path, keep: set[str], *, renumber_rank: bool) -> tuple[int, int]:
    """Filter ratingData / financeDynamic style files.

    These files have one JSON object per line between `[` and `];`.
    Returns (kept_count, removed_count).
    """
    text = path.read_text(encoding="utf-8")
    start_idx = text.find("= [\n")
    if start_idx == -1:
        sys.exit(f"{path.name}: cannot find '= [' array start")
    start_idx += len("= [\n")
    end_idx = text.rfind("\n];")
    if end_idx == -1:
        sys.exit(f"{path.name}: cannot find closing '];'")

    header = text[:start_idx]
    body = text[start_idx:end_idx]
    footer = text[end_idx:]

    lines = body.splitlines()
    kept_objects: list[dict] = []
    removed = 0

    for line in lines:
        m = OBJECT_LINE_RE.match(line)
        if not m:
            # blank or comment? skip silently
            if line.strip() == "" or line.strip().startswith("//"):
                continue
            sys.exit(f"{path.name}: unexpected line: {line!r}")
        obj = json.loads(m.group(1))
        inn = str(obj.get("inn", "")).strip()
        if inn in keep:
            kept_objects.append(obj)
        else:
            removed += 1

    # Preserve original sort order — but ranks might have gaps.
    # Sort by current rank asc so ties are resolved deterministically.
    kept_objects.sort(key=lambda o: o.get("rank", 10**9))

    if renumber_rank:
        for new_rank, obj in enumerate(kept_objects, start=1):
            obj["rank"] = new_rank

    # Serialize back, one object per line.
    # json.dumps without indent gives us the exact format already in the file.
    rendered_lines = []
    for obj in kept_objects:
        rendered = json.dumps(obj, ensure_ascii=False, separators=(", ", ": "))
        rendered_lines.append(f"  {rendered},")

    new_body = "\n".join(rendered_lines)
    new_text = header + new_body + footer

    path.write_text(new_text, encoding="utf-8")
    return len(kept_objects), removed


# ---------------------------------------------------------------------------
# Helpers for Record<INN, ...> files
# ---------------------------------------------------------------------------
RECORD_LINE_RE = re.compile(r'^\s*"(\d+)":\s*(.*?),?\s*$')


def filter_record_map(path: Path, keep: set[str], var_name: str) -> tuple[int, int]:
    """Filter a `export const <var_name>: Record<string, ...> = { ... };` file.

    Each line in the body is `"INN": <value>,`. Returns (kept, removed).
    """
    text = path.read_text(encoding="utf-8")
    open_marker = f"{var_name}"
    start = text.find(open_marker)
    if start == -1:
        sys.exit(f"{path.name}: '{var_name}' not found")
    brace_start = text.find("= {", start)
    if brace_start == -1:
        sys.exit(path.name + ": '= {' not found after " + var_name)
    body_start = brace_start + len("= {\n")
    body_end = text.rfind("\n};")
    if body_end == -1:
        sys.exit(path.name + ": closing '};' not found")

    header = text[:body_start]
    body = text[body_start:body_end]
    footer = text[body_end:]

    kept_lines: list[str] = []
    removed = 0
    kept_count = 0

    for raw in body.splitlines():
        stripped = raw.strip()
        if stripped == "" or stripped.startswith("//"):
            kept_lines.append(raw)
            continue
        m = RECORD_LINE_RE.match(raw)
        if not m:
            sys.exit(f"{path.name}: unexpected line: {raw!r}")
        inn = m.group(1)
        if inn in keep:
            # Ensure trailing comma for consistency.
            line = raw.rstrip()
            if not line.endswith(","):
                line += ","
            kept_lines.append(line)
            kept_count += 1
        else:
            removed += 1

    new_body = "\n".join(kept_lines)
    new_text = header + new_body + footer

    path.write_text(new_text, encoding="utf-8")
    return kept_count, removed


# ---------------------------------------------------------------------------
# investmentData verification (read-only)
# ---------------------------------------------------------------------------
def strip_org_form(name: str) -> str:
    s = name.strip()
    s = re.sub(r"^(НАО|ООО|ПАО|АО|ЗАО)\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r'^"?ПРОФЕССИОНАЛЬНАЯ КОЛЛЕКТОРСКАЯ ОРГАНИЗАЦИЯ\s*', "", s, flags=re.IGNORECASE)
    s = re.sub(r'^"?ПКО\s*', "", s, flags=re.IGNORECASE)
    s = re.sub(r'^["\s]+', "", s)
    s = re.sub(r'["\s]+$', "", s)
    return s.lower()


def verify_investment_data(kept_full_names: dict[str, str]) -> None:
    """For each `company: '...'` in investmentData.ts, verify its short name
    appears (case-insensitive substring) inside some kept fullName.
    Uses `fullName` from source JSON — authoritative and complete.
    """
    text = INVEST_TS.read_text(encoding="utf-8")
    inv_companies = set(re.findall(r"company:\s*'([^']+)'", text))

    # Normalize: lower + strip punctuation for loose substring match
    def norm(s: str) -> str:
        s = s.lower()
        s = re.sub(r"[«»\"'()\[\]\.,]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    kept_norm_names = [norm(n) for n in kept_full_names.values() if n]
    orphans: list[str] = []
    for inv_name in inv_companies:
        # Split on parentheses/slashes — "Brand (Legal)" → ["Brand", "Legal"]
        candidates = [c for c in re.split(r"[()\/]", inv_name) if c.strip()]
        if not candidates:
            candidates = [inv_name]
        matched = False
        for cand in candidates:
            cand_n = norm(cand)
            if not cand_n:
                continue
            if any(cand_n in kept for kept in kept_norm_names):
                matched = True
                break
        if not matched:
            orphans.append(inv_name)

    if orphans:
        print("\n⚠ investmentData.ts references companies not found in keep-list fullNames:")
        for inv in orphans:
            print(f"   '{inv}'")
        sys.exit("Aborted: resolve investmentData orphans manually.")
    else:
        print(f"✓ investmentData.ts: {len(inv_companies)} companies, all have a home")


# ---------------------------------------------------------------------------
# Extract name map from ratingData BEFORE pruning, so we can warn
# ---------------------------------------------------------------------------
def snapshot_inn_to_name(path: Path) -> dict[str, str]:
    text = path.read_text(encoding="utf-8")
    # Use a forgiving regex: capture inn, then walk back to the preceding "name"
    # Simpler: each object is on one line — parse each line as JSON.
    names: dict[str, str] = {}
    for line in text.splitlines():
        m = OBJECT_LINE_RE.match(line)
        if not m:
            continue
        try:
            obj = json.loads(m.group(1))
        except json.JSONDecodeError:
            continue
        if "inn" in obj and "name" in obj:
            names[str(obj["inn"])] = obj["name"]
    return names


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    keep, keep_full_names = load_keep_data()
    src = SOURCE_JSON.name if SOURCE_JSON.exists() else SOURCE_XLSX.name
    print(f"Keep list: {len(keep)} unique INNs from {src}")

    # Snapshot ratingData before mutation (to count what gets removed).
    all_names = snapshot_inn_to_name(RATING_TS)
    removed_map = {inn: name for inn, name in all_names.items() if inn not in keep}
    print(f"Current ratingData: {len(all_names)} entries, {len(removed_map)} to remove")

    # Verify investmentData using authoritative fullNames from source JSON.
    verify_investment_data(keep_full_names)

    # Mutate files.
    print()
    k, r = filter_line_array(RATING_TS, keep, renumber_rank=True)
    print(f"ratingData.ts:       kept {k}, removed {r}")

    k, r = filter_line_array(FINANCE_TS, keep, renumber_rank=True)
    print(f"financeDynamic.ts:   kept {k}, removed {r}")

    k, r = filter_record_map(DETAILS_TS, keep, "companyDetailsMap")
    print(f"companyDetails.ts:   kept {k}, removed {r}")

    k, r = filter_record_map(LOGO_TS, keep, "logoMap")
    print(f"logoMap.ts:          kept {k}, removed {r}")

    print("\nDone.")


if __name__ == "__main__":
    main()
